"""
============================
author:qideping
time:2021/8/17 2:51 下午
E-mail:qideping@makenv.com
============================
"""
import os
import openpyxl


class ReadExcl(object):

    def __init__(self, file_name, sheet_name):
        """

        :param file_name: excel的文件路径
        :param sheet_name: excel中sheet的名称
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.workbook = openpyxl.load_workbook(self.file_name)
        self.sheet = self.workbook[self.sheet_name]

    def create_data(self):
        self.open()  # 调用打开方法
        minrow = self.sheet.min_row  # 最小行
        maxrow = self.sheet.max_row  # 最大行
        mincol = self.sheet.min_column  # 最小列
        maxcol = self.sheet.max_column  # 最大列
        tb_1 = []
        # case_type = "点源文件上传-专用表-电力行业-电力企业-专用表_p101_机组信息表，"
        case_type = "点源文件上传-专用表_p101_机组信息表，"
        case_type1 = "{}验证【{}】所属排放源的设备类型可选项为{}"
        case_type1_1 = "{}验证【{}】所属排放源的设备类型填写非可选项{}报错校验"
        case_type2 = "{}验证【{}】所属排放源【{}】设备类型的燃料类型可选项为{}"
        case_type2_1 = "{}验证【{}】所属排放源【{}】设备类型的燃料类型填写非可选项{}报错校验"
        case_type3 = "{}验证【{}】所属排放源【{}】设备类型【{}】燃料类型的单位可选项为{}"
        case_type3_1 = "{}验证【{}】所属排放源【{}】设备类型【{}】燃料类型的单位填写非可选项{}报错校验"

        cases = []
        for n in range(minrow + 1, maxrow + 1):
            tb_1.append(self.sheet.cell(n, 1).value)

        for m in set(tb_1):
            tb_2 = []  # 存放所属排放源
            for n in range(minrow+1, maxrow+1):
                if self.sheet.cell(n, 1).value == m:
                    tb_2.append(self.sheet.cell(n, 2).value)
            case_title1 = case_type1.format(case_type, m, set(tb_2))
            case_title1_1 = case_type1_1.format(case_type, m, set(tb_2))
            case_1 = [case_title1, '正', 2]
            case_1_1 = [case_title1_1, '反', 2]
            # print(case_type1.format(case_type, m, set(tb_2)))
            # print(case_type1_1.format(case_type, m, set(tb_2)))
            cases.append(case_1)
            cases.append(case_1_1)
            for i in set(tb_2):
                tb_3 = []  # 存放产品类型
                for t in range(minrow + 1, maxrow + 1):
                    if self.sheet.cell(t, 2).value == i:
                        tb_3.append(self.sheet.cell(t, 3).value)
                # print(set(tb_3))
                # print(case_type2.format(case_type, m, i, set(tb_3)))

            for i in set(tb_2):
                tb_3 = []  # 存放产品类型
                for t in range(minrow + 1, maxrow + 1):
                    if self.sheet.cell(t, 2).value == i:
                        tb_3.append(self.sheet.cell(t, 3).value)
                # print(case_type2.format(case_type, m, i, set(tb_3)))
                for a in set(tb_3):
                    # print(a)
                    tb_4 = []  # 存放工段/工艺名称
                    for b in range(minrow + 1, maxrow + 1):
                        if self.sheet.cell(b, 3).value == a:
                            tb_4.append(self.sheet.cell(b, 4).value)
                    # print(set(tb_4))
                    # print(case_type3.format(case_type, m, i, a, set(tb_4)))

            for i in set(tb_2):
                tb_3 = []  # 存放产品类型
                for t in range(minrow + 1, maxrow + 1):
                    if self.sheet.cell(t, 2).value == i:
                        tb_3.append(self.sheet.cell(t, 3).value)
                case_title2 = case_type2.format(case_type, m, i, set(tb_3))
                case_title2_1 = case_type2_1.format(case_type, m, i, set(tb_3))
                case_2 = [case_title2, '正', 2]
                case_2_1 = [case_title2_1, '反', 3]
                # print(case_type2.format(case_type, m, i, set(tb_3)))
                # print(case_type2_1.format(case_type, m, i, set(tb_3)))
                cases.append(case_2)
                cases.append(case_2_1)
                for a in set(tb_3):
                    # print(a)
                    tb_4 = []  # 存放工段/工艺名称
                    for b in range(minrow + 1, maxrow + 1):
                        if self.sheet.cell(b, 3).value == a:
                            tb_4.append(self.sheet.cell(b, 4).value)
                    case_title3 = case_type3.format(case_type, m, i, a, set(tb_4))
                    case_title3_1 = case_type3_1.format(case_type, m, i, a, set(tb_4))
                    case_3 = [case_title3, '正', 2]
                    case_3_1 = [case_title3_1, '反', 3]
                    # print(case_type3.format(case_type, m, i, a, set(tb_4)))
                    # print(case_type3_1.format(case_type, m, i, a, set(tb_4)))
                    cases.append(case_3)
                    cases.append(case_3_1)
                    # for c in set(tb_4):
                    #     # print(c)
                    #     tb_5 = []  # 存放单位
                    #     for d in range(minrow + 1, maxrow + 1):
                    #         if self.sheet.cell(d, 4).value == c:
                    #             tb_5.append(self.sheet.cell(d, 5).value)
                    #     case_title4 = case_type4.format(case_type, m, i, a, c, set(tb_5))
                    #     case_title4_1 = case_type4_1.format(case_type, m, i, a, c, set(tb_5))
                    #     case_4 = [case_title4, '正', 2]
                    #     case_4_1 = [case_title4_1, '反', 3]
                    #     # print(case_type4.format(case_type, m, i, a, c, set(tb_5)))
                    #     # print(case_type4_1.format(case_type, m, i, a, c, set(tb_5)))
                    #     cases.append(case_4)
                    #     cases.append(case_4_1)
        return cases

    def write(self, row, colunm, value):
        # self.open()
        self.sheet.cell(row, colunm, value)
        # self.workbook.save(self.file_name)

    def save(self):
        self.workbook.save(self.file_name)


if __name__ == '__main__':
    data_path = "//create_test_cases/25.xlsx"
    case_path = "//create_test_cases/City项目-清单系统产品-级联关系校验.xlsx"
    print(data_path)
    r1 = ReadExcl(data_path, '机组')
    # r1 = ReadExcl(data_path, '专用表民用锅炉')
    cases = r1.create_data()
    r2 = ReadExcl(case_path, '专用表_p101_机组信息表')
    r2.open()
    for i in range(len(cases)):
        r2.write(i+2, 5, cases[i][0])
        r2.write(i+2, 10, cases[i][1])
        r2.write(i+2, 11, cases[i][2])
    r2.save()
